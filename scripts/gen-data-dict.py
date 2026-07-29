import sys, re, json, os, datetime
sys.path.insert(0, r'C:\Users\JFFC\.workbuddy\binaries\python\envs\default\Lib\site-packages')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = r'E:\project\NobleNest\Vshop'
SCHEMA_PATH = os.path.join(ROOT, 'vshop-server', 'prisma', 'schema.prisma')
SNAP_PATH = os.path.join(ROOT, '.workbuddy', 'schema-snapshot.json')
DOC_DIR = os.path.join(ROOT, 'docs')
MD_PATH = os.path.join(DOC_DIR, 'data-dictionary.md')
XLSX_PATH = os.path.join(DOC_DIR, 'data-dictionary.xlsx')
TODAY = datetime.date.today().strftime('%Y-%m-%d')

BASE_TYPES = {'String', 'Int', 'Float', 'Boolean', 'DateTime', 'Json', 'Bytes', 'Decimal', 'BigInt'}

FIELD_DESC = {
    'User.id': '用户唯一 ID', 'User.openid': '微信小程序 OpenID', 'User.unionId': '微信 UnionID（跨应用用户识别）',
    'User.nickname': '用户昵称', 'User.avatar': '头像 URL', 'User.phone': '手机号',
    'User.location': '用户位置（城市级别）', 'User.isKoc': '是否为 KOC 分销员', 'User.kocApprovedAt': 'KOC 审核通过时间',
    'User.createdAt': '创建时间', 'User.updatedAt': '更新时间',
    'Address.id': '地址 ID', 'Address.userId': '所属用户', 'Address.name': '收货人姓名',
    'Address.phone': '收货人手机号', 'Address.province': '省份', 'Address.city': '城市',
    'Address.district': '区/县', 'Address.detail': '详细地址', 'Address.isDefault': '是否默认地址',
    'Address.createdAt': '创建时间', 'Address.updatedAt': '更新时间',
    'Category.id': '分类 ID', 'Category.name': '分类名称', 'Category.icon': '分类图标 URL',
    'Category.sort': '排序（越小越前）', 'Category.status': '状态（active/disabled）',
    'SubCategory.id': '子分类 ID', 'SubCategory.categoryId': '所属一级分类', 'SubCategory.name': '子分类名称',
    'SubCategory.sort': '排序（越小越前）',
    'Good.id': '商品 ID', 'Good.subCategoryId': '所属子分类', 'Good.name': '商品名称',
    'Good.description': '商品简介', 'Good.detail': '商品详情（富文本）', 'Good.sales': '销量',
    'Good.status': '状态（active/offline）', 'Good.discountRate': '商品级折扣率（0~1，null 无折扣）',
    'Good.isRecommended': '是否今日推荐', 'Good.recommendSort': '推荐排序', 'Good.createdAt': '创建时间', 'Good.updatedAt': '更新时间',
    'GoodImage.id': '图片 ID', 'GoodImage.goodId': '所属商品', 'GoodImage.url': '图片 URL', 'GoodImage.sort': '排序',
    'GoodDetailImage.id': '图片 ID', 'GoodDetailImage.goodId': '所属商品', 'GoodDetailImage.url': '图片 URL', 'GoodDetailImage.sort': '排序（升序展示）',
    'Sku.id': 'SKU ID', 'Sku.goodId': '所属商品', 'Sku.name': 'SKU 名称（单规格回退展示）',
    'Sku.specValues': '规格维度组合 [{"name":"颜色","value":"红"}]', 'Sku.price': '售价（分）',
    'Sku.marketPrice': '市场价/划线价（分）', 'Sku.stock': '库存数量', 'Sku.platformSkuId': '店管家平台规格 ID',
    'Supplier.id': '供应商 ID', 'Supplier.name': '供应商名称', 'Supplier.contactName': '联系人',
    'Supplier.contactPhone': '联系电话', 'Supplier.deliveryRegions': '配送区域（城市 JSON 数组）',
    'Supplier.fulfillRate': '履约率（%）', 'Supplier.status': '状态（active/disabled）', 'Supplier.createdAt': '创建时间',
    'GoodSupplier.id': '关联 ID', 'GoodSupplier.goodId': '商品 ID', 'GoodSupplier.supplierId': '供应商 ID',
    'GoodSupplier.skuId': 'SKU ID', 'GoodSupplier.price': '供货价（分）', 'GoodSupplier.stock': '供货库存',
    'GoodSupplier.freight': '运费（分）', 'GoodSupplier.status': '状态',
    'CartItem.id': '购物车项 ID', 'CartItem.userId': '用户 ID', 'CartItem.skuId': 'SKU ID', 'CartItem.quantity': '数量',
    'Order.id': '订单 ID', 'Order.orderSn': '订单编号（展示用）', 'Order.userId': '用户 ID', 'Order.addressId': '收货地址 ID',
    'Order.supplierId': '供应商 ID', 'Order.status': '订单状态（pending/shipping/receiving/done）', 'Order.totalAmount': '商品总金额（分）',
    'Order.discountAmount': '优惠金额（分）', 'Order.freightAmount': '运费（分）', 'Order.payAmount': '实付金额（分）',
    'Order.remark': '订单备注', 'Order.platformOrderId': '店管家平台单号', 'Order.dianjiaSyncStatus': '店管家同步状态（synced/failed）',
    'Order.dianjiaSyncedAt': '最近同步时间', 'Order.kocId': 'KOC 分销员 ID（分销佣金溯源）',
    'Order.createdAt': '创建时间', 'Order.updatedAt': '更新时间',
    'OrderItem.id': '订单项 ID', 'OrderItem.orderId': '订单 ID', 'OrderItem.skuId': 'SKU ID',
    'OrderItem.goodTitle': '商品标题（下单快照）', 'OrderItem.specName': '规格名称（下单快照）',
    'OrderItem.image': '商品图片（下单快照）', 'OrderItem.price': '下单单价（分）', 'OrderItem.quantity': '数量',
    'OrderPackage.id': '包裹 ID', 'OrderPackage.orderId': '订单 ID', 'OrderPackage.supplierId': '供应商 ID',
    'OrderPackage.supplierName': '供应商名称（快照）', 'OrderPackage.status': '物流状态（0待发货 1已打单 2已发货 3运输中 4派送中 5已签收）',
    'OrderPackage.expressCompany': '快递公司', 'OrderPackage.expressNo': '快递单号',
    'Payment.id': '支付 ID', 'Payment.orderId': '订单 ID（1:1）', 'Payment.amount': '支付金额（分）',
    'Payment.status': '支付状态（pending/success/failed/refund）', 'Payment.payTime': '支付时间',
    'Payment.transactionId': '微信支付交易号', 'Payment.prepayId': '微信预支付 ID', 'Payment.provider': '支付渠道（wechatpay/mock）',
    'Payment.createdAt': '创建时间',
    'Favorite.id': '收藏 ID', 'Favorite.userId': '用户 ID', 'Favorite.goodId': '商品 ID', 'Favorite.createdAt': '收藏时间',
    'Coupon.id': '优惠券模板 ID', 'Coupon.name': '优惠券名称', 'Coupon.type': '类型（cash 现金券 / discount 折扣券）',
    'Coupon.value': '面额（分），type=cash 使用', 'Coupon.discountValue': '折扣率（0~1），type=discount 使用',
    'Coupon.minAmount': '满减门槛（分）', 'Coupon.scopeType': '适用范围（all 全场）', 'Coupon.totalCount': '发放总量',
    'Coupon.usedCount': '已领取数', 'Coupon.expireTime': '过期时间', 'Coupon.status': '状态（active/disabled）',
    'UserCoupon.id': '领取记录 ID', 'UserCoupon.userId': '用户 ID', 'UserCoupon.couponId': '优惠券模板 ID',
    'UserCoupon.status': '状态（usable/used/expired）', 'UserCoupon.usedAt': '使用时间', 'UserCoupon.createdAt': '领取时间',
    'ChannelReport.id': '归因记录 ID', 'ChannelReport.source': '来源渠道（koc/card）',
    'ChannelReport.kocId': 'KOC ID（source=koc 时）', 'ChannelReport.batchId': '批次 ID（包裹卡片批号）', 'ChannelReport.createdAt': '记录时间',
    'KocProfile.id': '资料 ID', 'KocProfile.userId': '用户 ID（1:1）', 'KocProfile.realName': '真实姓名',
    'KocProfile.phone': '手机号', 'KocProfile.socialAccount': '社交账号（微信号/手机）', 'KocProfile.introduction': '自我介绍/申请说明',
    'KocProfile.status': '审核状态（pending/approved/rejected/disabled）', 'KocProfile.rejectReason': '驳回原因',
    'KocProfile.commissionRate': '自定义佣金率（0~1），null 走系统默认阶梯', 'KocProfile.reviewedAt': '审核时间',
    'KocProfile.createdAt': '申请时间', 'KocProfile.updatedAt': '更新时间',
    'ChatSession.id': '会话 ID', 'ChatSession.userId': '用户 ID', 'ChatSession.goodId': '关联商品（从商品页发起时）',
    'ChatSession.title': '会话标题', 'ChatSession.lastMessage': '最后一条消息（冗余）', 'ChatSession.lastAt': '最后消息时间',
    'ChatSession.userUnread': '用户未读数', 'ChatSession.adminUnread': '客服未读数', 'ChatSession.closed': '是否已关闭',
    'ChatSession.createdAt': '创建时间', 'ChatSession.updatedAt': '更新时间',
    'ChatMessage.id': '消息 ID', 'ChatMessage.sessionId': '会话 ID', 'ChatMessage.sender': '发送者（user/admin）',
    'ChatMessage.content': '消息内容', 'ChatMessage.createdAt': '发送时间',
    'Aftersale.id': '售后 ID', 'Aftersale.aftersaleSn': '售后单号', 'Aftersale.userId': '用户 ID',
    'Aftersale.orderId': '订单 ID', 'Aftersale.orderItemId': '售后商品项 ID', 'Aftersale.packageIndex': '包裹序号',
    'Aftersale.type': '售后类型（1 仅退款 / 2 退货退款）', 'Aftersale.reason': '售后原因', 'Aftersale.description': '问题描述',
    'Aftersale.evidenceImages': '凭证图片 URL 数组', 'Aftersale.refundAmount': '退款金额（分）',
    'Aftersale.refundNo': '退款单号（out_refund_no）', 'Aftersale.refundId': '微信退款单号',
    'Aftersale.refundStatus': '微信退款状态（PROCESSING/SUCCESS/CLOSED/ABNORMAL）', 'Aftersale.adminRemark': '管理员备注',
    'Aftersale.status': '审核状态（0待审核 1已同意 2已拒绝 3退款中 4已退款 5已完成）',
    'Aftersale.createdAt': '创建时间', 'Aftersale.updatedAt': '更新时间',
    'AppSetting.key': '配置键（如 dianjia_auto_sync）', 'AppSetting.value': '配置值', 'AppSetting.updatedAt': '更新时间',
}

DESC_MAP = {
    'User': '微信小程序用户基础信息 + KOC 标识',
    'Address': '用户收货地址，支持多地址、默认地址',
    'Category': '一级分类（如水果、蔬菜、肉类）',
    'SubCategory': '二级分类，归属一级分类',
    'Good': '商品主表，含折扣、推荐标记',
    'GoodImage': '商品主图（轮播图），按 sort 排序',
    'GoodDetailImage': '详情页纵向铺图，按 sort 排序',
    'Sku': '多规格库存单位，价格以分为单位',
    'Supplier': '供应商基础信息与履约数据',
    'GoodSupplier': '商品-SKU-供应商三方关联（含供货价/库存/运费）',
    'CartItem': '用户购物车，按用户+SKU 去重',
    'Order': '订单主表，含店管家同步状态',
    'OrderItem': '订单内商品明细（冗余商品标题/规格/图片）',
    'OrderPackage': '订单包裹物流信息（一单多包场景）',
    'Payment': '微信支付记录，与订单 1:1',
    'Favorite': '用户商品收藏，按用户+商品去重',
    'Coupon': '优惠券定义（现金券/折扣券），支持库存管理',
    'UserCoupon': '用户领取的优惠券实例与使用状态',
    'ChannelReport': '用户来源渠道埋点（KOC/包裹卡片）',
    'KocProfile': 'KOC 申请、审核、佣金配置',
    'ChatSession': '在线客服会话记录，含未读计数',
    'ChatMessage': '会话内消息明细',
    'Aftersale': '退款/退货退款申请与处理',
    'AppSetting': '通用 key-value 配置（如自动同步开关）',
}

# ============================================================
# 1. Parse Prisma schema
# ============================================================
def parse_prisma(text):
    models = {}
    # find model blocks
    model_blocks = re.findall(r'model\s+(\w+)\s*\{([^}]*)\}', text, re.DOTALL)
    # also capture enums if any (skip for now)
    for mname, body in model_blocks:
        fields = []
        indexes = []
        relations = []
        for line in body.split('\n'):
            line = line.strip()
            if not line or line.startswith('//'):
                continue
            # index / unique
            m_idx = re.match(r'@@(index|unique)\((.*)\)', line)
            if m_idx:
                kind = m_idx.group(1)
                args = m_idx.group(2)
                # extract field names inside [..]
                fm = re.search(r'\[(.*?)\]', args)
                flds = [x.strip() for x in fm.group(1).split(',')] if fm else []
                indexes.append({'type': kind, 'fields': flds})
                continue
            # skip relation attribute lines like @@relation
            if line.startswith('@@'):
                continue
            # field line
            fm = re.match(r'(\w+)\s+(\w+)(\[\])?(\?)?\s*(.*)', line)
            if not fm:
                continue
            fname = fm.group(1)
            ftype = fm.group(2)
            is_list = bool(fm.group(3))
            is_opt = bool(fm.group(4))
            rest = fm.group(5)
            is_id = '@id' in rest
            is_unique = '@unique' in rest
            is_updated = '@updatedAt' in rest
            dm = re.search(r'@default\((.*)\)\s*$', rest)
            default = dm.group(1) if dm else None
            # relation?
            rel = None
            if ftype not in BASE_TYPES:
                card = 'many' if is_list else ('one' if is_opt else 'one')
                rel = {'model': ftype, 'cardinality': card}
                relations.append(rel)
            fields.append({
                'name': fname, 'type': ftype, 'isList': is_list, 'isOptional': is_opt,
                'isId': is_id, 'isUnique': is_unique, 'isUpdatedAt': is_updated,
                'default': default, 'relation': rel,
            })
        models[mname] = {'name': mname, 'fields': fields, 'indexes': indexes,
                        'relations': relations, 'desc': DESC_MAP.get(mname, '')}
    return models

def field_desc(table, fname):
    return FIELD_DESC.get(f"{table}.{fname}", '')

# ============================================================
# 2. Field signature for diff
# ============================================================
def field_constraint_sig(f):
    parts = []
    parts.append('必填' if not f['isOptional'] else '可空')
    if f['isId']:
        parts.append('PK')
    if f['isUnique']:
        parts.append('唯一')
    if f['isUpdatedAt']:
        parts.append('auto')
    return ' / '.join(parts)

def field_full_sig(f):
    d = f['default'] or '—'
    return f"{f['type']} | {field_constraint_sig(f)} | 默认:{d}"

def model_sig(m):
    return {
        'fields': {f['name']: field_full_sig(f) for f in m['fields'] if not f['relation']},
        'indexes': sorted([f"{i['type']}:{','.join(i['fields'])}" for i in m['indexes']]),
        'relations': sorted([f"{r['model']}:{r['cardinality']}" for r in m['relations']]),
    }

# ============================================================
# 3. Diff
# ============================================================
def diff_models(old, new):
    changes = []
    old_names = set(old.keys())
    new_names = set(new.keys())
    for name in sorted(new_names - old_names):
        changes.append({'type': '新增表', 'table': name, 'field': '—',
                        'before': '—', 'after': '新建数据表', 'note': '新建表'})
    for name in sorted(old_names - new_names):
        changes.append({'type': '删除表', 'table': name, 'field': '—',
                        'before': '存在', 'after': '已删除', 'note': '删除表'})
    for name in sorted(old_names & new_names):
        om, nm = old[name], new[name]
        osig, nsig = model_sig(om), model_sig(nm)
        # field adds/deletes/modifies
        of, nf = osig['fields'], nsig['fields']
        for fn in sorted(set(nf) - set(of)):
            changes.append({'type': '字段新增', 'table': name, 'field': fn,
                            'before': '—', 'after': nf[fn], 'note': '新增字段'})
        for fn in sorted(set(of) - set(nf)):
            changes.append({'type': '字段删除', 'table': name, 'field': fn,
                            'before': of[fn], 'after': '—', 'note': '删除字段'})
        for fn in sorted(set(of) & set(nf)):
            if of[fn] != nf[fn]:
                changes.append({'type': '字段修改', 'table': name, 'field': fn,
                                'before': of[fn], 'after': nf[fn], 'note': '字段定义变更'})
        # indexes
        oi, ni = set(osig['indexes']), set(nsig['indexes'])
        for ix in sorted(ni - oi):
            changes.append({'type': '索引新增', 'table': name, 'field': '—',
                            'before': '—', 'after': ix, 'note': '新增索引'})
        for ix in sorted(oi - ni):
            changes.append({'type': '索引删除', 'table': name, 'field': '—',
                            'before': ix, 'after': '—', 'note': '删除索引'})
        # relations
        orl, nrl = set(osig['relations']), set(nsig['relations'])
        for rl in sorted(nrl - orl):
            changes.append({'type': '关系新增', 'table': name, 'field': '—',
                            'before': '—', 'after': rl, 'note': '新增关联关系'})
        for rl in sorted(orl - nrl):
            changes.append({'type': '关系删除', 'table': name, 'field': '—',
                            'before': rl, 'after': '—', 'note': '删除关联关系'})
    return changes

# ============================================================
# 4. Load / save snapshot + version
# ============================================================
parsed = parse_prisma(open(SCHEMA_PATH, encoding='utf-8').read())
if os.path.exists(SNAP_PATH):
    snap = json.load(open(SNAP_PATH, encoding='utf-8'))
    old_models = snap['models']
    old_version = snap['version']
    has_prev = True
else:
    old_models = {}
    old_version = 'v1.0'
    has_prev = False

# build snapshot models (full parsed models for accurate diff)
snap_models = {name: m for name, m in parsed.items()}

changes = diff_models(old_models, parsed) if has_prev else []

if has_prev:
    if changes:
        # bump minor
        try:
            major, minor = old_version[1:].split('.')
            new_version = f"v{major}.{int(minor)+1}"
        except Exception:
            new_version = f"{old_version}+1"
        changed = True
    else:
        new_version = old_version
        changed = False
else:
    new_version = 'v1.0'
    changed = True  # initial baseline

# load existing changelog from snapshot
changelog = snap.get('changelog', []) if has_prev else []

if changed:
    batch = {
        'version': new_version,
        'date': TODAY,
        'summary': '初始化数据字典基线' if not has_prev else
                   ('无结构变更' if not changes else f'共 {len(changes)} 项变更'),
        'changes': changes if changes else [],
    }
    changelog.append(batch)

# save snapshot
json.dump({'version': new_version, 'date': TODAY, 'models': snap_models, 'changelog': changelog},
          open(SNAP_PATH, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

# ============================================================
# 5. Markdown generation
# ============================================================
def md_escape(s):
    return s.replace('|', '\\|')

def cn_name(name):
    cns = {
        'User':'用户表','Address':'收货地址表','Category':'商品分类表（一级）','SubCategory':'商品子分类表（二级）',
        'Good':'商品表','GoodImage':'商品图片表（轮播图）','GoodDetailImage':'商品详情图表','Sku':'商品 SKU 表',
        'Supplier':'供应商表','GoodSupplier':'商品供应商关联表','CartItem':'购物车表','Order':'订单表',
        'OrderItem':'订单商品项表','OrderPackage':'订单包裹表','Payment':'支付表','Favorite':'收藏表',
        'Coupon':'优惠券模板表','UserCoupon':'用户优惠券表','ChannelReport':'渠道归因表','KocProfile':'KOC 分销员资料表',
        'ChatSession':'客服会话表','ChatMessage':'客服消息表','Aftersale':'售后表','AppSetting':'应用设置表',
    }
    return cns.get(name, name)

md = []
md.append('# Vshop 数据库数据字典\n')
md.append(f'> 基于 `vshop-server/prisma/schema.prisma` 生成  \n> 数据库类型：MySQL | ORM：Prisma | 当前版本：{new_version} | 更新时间：{TODAY}\n')
md.append('---\n')

# changelog section first
md.append('## 变更记录\n')
if not changelog:
    md.append('暂无变更记录。\n')
else:
    md.append('| 版本 | 日期 | 变更类型 | 表名 | 字段 | 变更前 | 变更后 | 说明 |')
    md.append('|------|------|----------|------|------|--------|--------|------|')
    for b in reversed(changelog):
        if not b['changes']:
            md.append(f"| {b['version']} | {b['date']} | — | — | — | — | — | {md_escape(b['summary'])} |")
        else:
            for c in b['changes']:
                md.append(f"| {b['version']} | {b['date']} | {c['type']} | {c['table']} | {c['field']} | {md_escape(c['before'])} | {md_escape(c['after'])} | {md_escape(c['note'])} |")
md.append('\n---\n')

md.append('## 表概览\n')
md.append('| 序号 | 表名 | 中文名 | 字段数 | 说明 |')
md.append('|------|------|--------|--------|------|')
for i, (name, m) in enumerate(sorted(parsed.items()), 1):
    nfields = len([f for f in m['fields'] if not f['relation']])
    md.append(f"| {i} | {name} | {cn_name(name)} | {nfields} | {m.get('desc','')} |")
md.append('\n---\n')

for i, (name, m) in enumerate(sorted(parsed.items()), 1):
    md.append(f'## {i}. {name} — {cn_name(name)}\n')
    md.append('| 字段 | 类型 | 必填 | 约束 | 默认值 | 说明 |')
    md.append('|------|------|------|------|--------|------|')
    for f in m['fields']:
        if f['relation']:
            continue
        req = '✅' if not f['isOptional'] else '—'
        cons = []
        if f['isId']: cons.append('PK')
        if f['isUnique']: cons.append('唯一')
        if f['relation']: cons.append('FK')
        if f['isUpdatedAt']: cons.append('auto')
        cons_str = ' / '.join(cons) if cons else '—'
        default = f['default'] or '—'
        desc = field_desc(name, f['name'])
        md.append(f"| {f['name']} | {f['type']} | {req} | {cons_str} | {md_escape(default)} | {md_escape(desc)} |")
    if m['indexes']:
        md.append('\n**索引：**')
        for ix in m['indexes']:
            md.append(f"- `@@{ix['type']}([{','.join(ix['fields'])}])`")
    if m['relations']:
        md.append('\n**关联关系：**')
        for r in m['relations']:
            md.append(f"- `{r['model']}`（{r['cardinality']}）")
    md.append('\n')

open(MD_PATH, 'w', encoding='utf-8').write('\n'.join(md))
print(f"Markdown written: {MD_PATH}")

# ============================================================
# 6. Excel generation
# ============================================================
wb = Workbook()

HEADER_FILL = PatternFill('solid', start_color='2F5496')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, color='2F5496', size=14)
SUBTITLE_FONT = Font(name='Arial', bold=True, color='333333', size=11)
BODY_FONT = Font(name='Arial', size=10, color='333333')
PK_FILL = PatternFill('solid', start_color='FFF2CC')
FK_FILL = PatternFill('solid', start_color='E2EFDA')
IDX_FILL = PatternFill('solid', start_color='D6E4F0')
REL_FILL = PatternFill('solid', start_color='FCE4D6')
ALT_FILL = PatternFill('solid', start_color='F5F5F5')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
BORDER = Border(left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
                top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7'))

def style_header(ws, row, ncols):
    for col in range(1, ncols+1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER

# --- 概览 ---
ws = wb.active; ws.title = '概览'
ws['A1'] = 'Vshop 数据库数据字典'; ws['A1'].font = Font(name='Arial', bold=True, color='2F5496', size=16)
ws.merge_cells('A1:E1'); ws['A1'].alignment = CENTER
ws['A2'] = f'数据库类型: MySQL | ORM: Prisma | 当前版本: {new_version} | 更新时间: {TODAY}'
ws['A2'].font = Font(name='Arial', size=10, color='666666'); ws.merge_cells('A2:E2')
hdr = ['序号','表名','中文名','字段数','说明']
for col, h in enumerate(hdr, 1):
    ws.cell(row=4, column=col, value=h)
style_header(ws, 4, len(hdr))
for i, (name, m) in enumerate(sorted(parsed.items()), 1):
    nfields = len([f for f in m['fields'] if not f['relation']])
    r = 4 + i
    for col, val in enumerate([i, name, cn_name(name), nfields, m.get('desc','')], 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = BODY_FONT
        c.alignment = LEFT if col in (2,3,5) else CENTER
        c.border = BORDER
        if i % 2 == 0:
            c.fill = ALT_FILL
ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 50
ws.freeze_panes = 'A5'

# --- 变更记录 ---
ws = wb.create_sheet('变更记录')
ws['A1'] = '变更记录'; ws['A1'].font = TITLE_FONT; ws.merge_cells('A1:H1')
hdr = ['版本','日期','变更类型','表名','字段','变更前','变更后','说明']
for col, h in enumerate(hdr, 1):
    ws.cell(row=3, column=col, value=h)
style_header(ws, 3, len(hdr))
r = 4
if changelog:
    for b in reversed(changelog):
        if not b['changes']:
            vals = [b['version'], b['date'], '—','—','—','—','—', b['summary']]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val); c.font = BODY_FONT
                c.alignment = LEFT if col == 8 else CENTER; c.border = BORDER
            r += 1
        else:
            for ch in b['changes']:
                vals = [b['version'], b['date'], ch['type'], ch['table'], ch['field'], ch['before'], ch['after'], ch['note']]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=r, column=col, value=val); c.font = BODY_FONT
                    c.alignment = LEFT if col in (3,4,5,6,7,8) else CENTER; c.border = BORDER
                    # color by type
                    if col == 3:
                        if '新增' in ch['type']: c.fill = FK_FILL
                        elif '删除' in ch['type']: c.fill = REL_FILL
                        elif '修改' in ch['type']: c.fill = IDX_FILL
                r += 1
else:
    ws.cell(row=4, column=1, value='暂无变更记录')
ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12; ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16; ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 30; ws.column_dimensions['H'].width = 20
ws.freeze_panes = 'A4'

# --- per-table sheets ---
for name, m in sorted(parsed.items()):
    ws = wb.create_sheet(title=name[:31])
    ws['A1'] = f'{name} — {cn_name(name)}'; ws['A1'].font = TITLE_FONT; ws.merge_cells('A1:F1')
    ws['A2'] = f"说明: {m.get('desc','')}"; ws['A2'].font = Font(name='Arial', size=10, color='666666'); ws.merge_cells('A2:F2')
    hdr = ['字段名','类型','必填','约束','默认值','说明']
    for col, h in enumerate(hdr, 1):
        ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, len(hdr))
    r = 5
    for f in m['fields']:
        if f['relation']:
            continue
        req = '是' if not f['isOptional'] else '否'
        cons = []
        if f['isId']: cons.append('PK')
        if f['isUnique']: cons.append('唯一')
        if f['relation']: cons.append('FK')
        if f['isUpdatedAt']: cons.append('auto')
        cons_str = ' / '.join(cons) if cons else '—'
        default = f['default'] or '—'
        vals = [f['name'], f['type'], req, cons_str, default, field_desc(name, f['name'])]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val); c.font = BODY_FONT
            c.alignment = CENTER if col in (2,3,4,5) else LEFT; c.border = BORDER
            if cons_str == 'PK': c.fill = PK_FILL
            elif 'FK' in cons_str: c.fill = FK_FILL
            elif '唯一' in cons_str: c.fill = IDX_FILL
            elif (r % 2 == 1): c.fill = ALT_FILL
        r += 1
    r += 1
    if m['indexes']:
        ws.cell(row=r, column=1, value='索引:').font = SUBTITLE_FONT; r += 1
        for ix in m['indexes']:
            c = ws.cell(row=r, column=1, value=f"@@{ix['type']}([{','.join(ix['fields'])}])")
            c.font = BODY_FONT; c.fill = IDX_FILL; c.border = BORDER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
        r += 1
    if m['relations']:
        ws.cell(row=r, column=1, value='关联关系:').font = SUBTITLE_FONT; r += 1
        for rl in m['relations']:
            c = ws.cell(row=r, column=1, value=f"{rl['model']}（{rl['cardinality']}）")
            c.font = BODY_FONT; c.fill = REL_FILL; c.border = BORDER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 8; ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16; ws.column_dimensions['F'].width = 48
    ws.freeze_panes = 'A5'

# --- 枚举值 ---
ws = wb.create_sheet('枚举值汇总')
ws['A1'] = '枚举值汇总'; ws['A1'].font = TITLE_FONT; ws.merge_cells('A1:B1')
ENUM_MAP = {
    'Order.status': [['pending','待支付'],['shipping','配送中'],['receiving','待收货'],['done','已完成']],
    'Payment.status': [['pending','待支付'],['success','支付成功'],['failed','支付失败'],['refund','已退款']],
    'KocProfile.status': [['pending','待审核'],['approved','已通过'],['rejected','已驳回'],['disabled','后台停用']],
    'Coupon.type': [['cash','现金券（固定面额）'],['discount','折扣券（百分比折扣）']],
    'UserCoupon.status': [['usable','可用'],['used','已使用'],['expired','已过期']],
    'Aftersale.type': [['1','仅退款'],['2','退货退款']],
    'Aftersale.status': [['0','待审核'],['1','已同意'],['2','已拒绝'],['3','退款中'],['4','已退款'],['5','已完成']],
    'OrderPackage.status': [['0','待发货'],['1','已打单'],['2','已发货'],['3','运输中'],['4','派送中'],['5','已签收']],
    'ChannelReport.source': [['koc','KOC 分销'],['card','包裹卡片']],
}
r = 3
for title, items in ENUM_MAP.items():
    ws.cell(row=r, column=1, value=title).font = SUBTITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 1
    for col, h in enumerate(['值','说明'], 1):
        c = ws.cell(row=r, column=col, value=h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
    r += 1
    for it in items:
        for col, val in enumerate(it, 1):
            c = ws.cell(row=r, column=col, value=val); c.font = BODY_FONT; c.alignment = LEFT if col==2 else CENTER; c.border = BORDER
        r += 1
    r += 1
ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 35

# --- 金额字段 ---
ws = wb.create_sheet('金额字段约定')
ws['A1'] = '金额字段约定'; ws['A1'].font = TITLE_FONT; ws.merge_cells('A1:B1')
ws['A2'] = '所有金额字段均以「分（Int）」为单位存储，避免浮点累积误差。Service 层通过 centToYuan / yuanToCent 转换为「元」对外输出。'
ws['A2'].font = Font(name='Arial', size=10, color='666666'); ws.merge_cells('A2:B2'); ws['A2'].alignment = LEFT
AMOUNT = [['Sku','price, marketPrice'],['GoodSupplier','price, freight'],['Order','totalAmount, discountAmount, freightAmount, payAmount'],
           ['OrderItem','price'],['Payment','amount'],['Coupon','value, minAmount'],['Aftersale','refundAmount']]
for col, h in enumerate(['表名','金额字段'], 1):
    ws.cell(row=4, column=col, value=h)
style_header(ws, 4, 2)
for i, (t, f) in enumerate(AMOUNT):
    r = 5 + i
    for col, val in enumerate([t, f], 1):
        c = ws.cell(row=r, column=col, value=val); c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
        if i % 2 == 1: c.fill = ALT_FILL
ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 55

wb.save(XLSX_PATH)
print(f"Excel written: {XLSX_PATH}")
print(f"Version: {new_version} | Changes: {len(changes)}")
